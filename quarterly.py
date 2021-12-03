# Import OpenPyXl and Datetime
import openpyxl
import datetime

# Setup a list for each office.
txk_list = []; paris_list = []; mtp_list = []; ss_list = []

ss = 0; paris = 0; txk = 0; mtp = 0
# Function creates new tab called month and outputs survey data one month or less old.
def ninetydaysold():
    wb = openpyxl.load_workbook("C:\\Users\\short\\Desktop\\JobSeek1.xlsx")
    wb.create_sheet(index=3, title='Quarterly')
    mtab = wb['Quarterly']
    mtab['A1'] = 'Total Past Quarter Office Counts'
    mtab['A2'] = 'Surveys from the past 90 days.'
    mtab['B2'] = onequart
    mtab['A3'] = 'Total of all surveys to date.'
    mtab['B3'] = morequart + onequart
    mtab['A5'] = 'Office count for SS is:'
    mtab['B5'] = ss
    mtab['A6'] = 'Office count for Paris is: '
    mtab['B6'] = paris
    mtab['A7'] = 'Office count for Texarkana is: '
    mtab['B7'] = txk
    mtab['A8'] = 'Office count for Mt Pleasant is: '
    mtab['B8'] = mtp
    wb.save("C:\\Users\\short\\Desktop\\JobSeek1.xlsx")

# Function counts how many surveys are less than 30 days old and how many are older than 30 days old
def checkdate():
    wb = openpyxl.load_workbook("C:\\Users\\short\\Desktop\\JobSeek1.xlsx")
    ws = wb['Data']
    # Initailize global variables
    global onequart; global morequart
    # Set Global variables to 0
    onequart = 0; morequart = 0
    # Initialize monthold to 30 days ago
    quarterold = datetime.datetime.now() - datetime.timedelta(days=90)
    print(quarterold)
    try:
        for cell in ws['A']:
            if cell.value == 'Start Date':
                print('skipping first row')
            elif cell.value > quarterold:
                onequart += 1
                print("not too old")
            else:
                morequart += 1
                print("too old")
        wb.save("C:\\Users\\short\\Desktop\\JobSeek1.xlsx")
    except:
        print("Found the end of check data!")
    else:
        print("Success in Check Date.")

# Function to see if survey is too old to count in the one month tally.        
def office_check():
    wb = openpyxl.load_workbook("C:\\Users\\short\\Desktop\\JobSeek1.xlsx")
    source = wb["Data"]
    # Initialize rows and colums to start in
    x = 2; y = 1; xr = 2; yc = 5
    quarterold = datetime.datetime.now() - datetime.timedelta(days=90)
    # Initialize global variables
    global ss; global paris; global txk; global mtp
    # Initailize global variables to 0
    ss = 0; paris = 0; txk = 0; mtp = 0
    try:
        for cell in source['A']:
            if source.cell(row=x, column=y).value > quarterold:
                txk_list.append(source.cell(row=xr, column=yc).value)
                x += 1
                xr += 1
            elif sorce.cell.value == None:
                print("Empty Cell")
            else:
                print('Too Old.')
                x += 1
                xr += 1
        for cell in source['C']:
            paris_list.append(cell.value)
        for cell in source['B']:
            mtp_list.append(cell.value)
        for cell in source['D']:
            ss_list.append(cell.value)
        for x in txk_list:
            if x == 'Texarkana':
                txk += 1
        for x in mtp_list:
            if x == 'Mount Pleasant':
                mtp += 1
        for x in ss_list:
            if x == 'Sulphur Springs':
                ss += 1
        for x in paris_list:
            if x == 'Paris':
                paris += 1
        wb.save("C:\\Users\\short\\Desktop\\JobSeek1.xlsx")
    except:
        print("Found the end of the survey!")
    else:
        print("Success in Office Check.")

# Main execution
if __name__ == '__main__':
    checkdate()
    office_check()
    ninetydaysold()
    print('Surveys within the past 30 days', onequart)
    print('Surveys that are more than 30 days old', morequart)
    print('Total Surveys', onequart + morequart)