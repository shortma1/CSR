# This script tallys data for the Year to Date portion of the JobSeeker survey for the Customer Service Report.
# Import OpenPyXl
from openpyxl import load_workbook
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils import datetime

# Initialize office list
txk_list = []; paris_list = []; mtp_list = []; ss_list = []

# Initialize question list
question1 = []; question2 = []; question3 = []; question4 = []; question5 = []; question6 = []; question7 = []; question8 = []; question9 = []

# Function Deletes extra stuff not needed in original spreadsheet and create new spreadsheet and puts the new data in it.
def del_extra_stuff():
    try:
        wb = load_workbook("C:\\Users\\short\\Desktop\\JobSeek.xlsx")
        ws = wb.active
        ws.delete_cols(1, 2)
        ws.delete_rows(2)
        ws.delete_cols(2,13)
        ws.delete_cols(6, 11)
        ws.delete_cols(9)
        ws.delete_cols(10)
        ws.delete_cols(11)
        ws.delete_cols(12)
        ws.delete_cols(13)
        ws.delete_cols(14)
        ws.delete_cols(15, 6)
        ws.title = 'Data'
        wb.save("C:\\Users\\short\\Desktop\\JobSeek1.xlsx")
    except:
        print("An error occured deleting extra stuff from spreadsheet!")
    else:
        print("Success deleting extra stuff.")

# Function counts how many surveys filled out for each individual office.
def count_offices():
    # Initialize global variables
    global txk; global mtp; global ss; global paris
    # Set global variables to 0
    txk = 0; ss = 0; mtp = 0; paris = 0
    try:
        wb = load_workbook("C:\\Users\\short\\Desktop\\JobSeek1.xlsx")
        source = wb["Data"]
        for cell in source['E']:
            txk_list.append(cell.value)
        for cell in source['C']:
            paris_list.append(cell.value)
        for cell in source['B']:
            mtp_list.append(cell.value)
        for cell in source['D']:
            ss_list.append(cell.value) 
        for x in txk_list:
            if x == 'Texarkana':
                txk += 1
        for x in mtp_list:
            if x == 'Mount Pleasant':
                mtp += 1
        for x in ss_list:
            if x == 'Sulphur Springs':
                ss += 1
        for x in paris_list:
            if x == 'Paris':
                paris += 1
        wb.save("C:\\Users\\short\\Desktop\\JobSeek1.xlsx")
    except:
        print("An error occured while counting offices!")
    else:
        print("Success while counting offices.")

# Function tally's up all of the questions
def question_tally():
    try:
        # Initialize global question variables
        global q1; global q2; global q3; global q4; global q5; global q6; global q7; global q8; global q9
        # Initialize question variables to 0
        q1 = 0; q2 = 0; q3 = 0; q4 = 0; q5 = 0; q6 = 0; q7 = 0; q8 = 0; q9 = 0
        wb = load_workbook("C:\\Users\\short\\Desktop\\JobSeek1.xlsx")
        source = wb["Data"]
        for cell in source['F']:
            question1.append(cell.value)
        for cell in source['G']:
            question2.append(cell.value)
        for cell in source['H']:
            question3.append(cell.value)
        for cell in source['I']:
            question4.append(cell.value)        
        for cell in source['J']:
            question5.append(cell.value)
        for cell in source['K']:
            question6.append(cell.value)
        for cell in source['L']:
            question7.append(cell.value)
        for cell in source['M']:
            question8.append(cell.value) 
        for cell in source['N']:
            question9.append(cell.value)
        for x in question1:
            if x == 'Yes':
                q1 += 1
            elif x == 'I did not need help with resources in the Center':
                q1 += 1
        for x in question2:
            if x == 'Yes':
                q2 += 1
            elif x == 'I did not need help with resources in the Center':
                q2 += 1
        for x in question3:
            if x == 'Agree':
                q3 += 1
            elif x == 'Strongly Agree':
                q3 += 1
        for x in question4:
            if x == 'Agree':
                q4 += 1
            elif x == 'Strongly Agree':
                q4 += 1
        for x in question5:
            if x == 'Agree':
                q5 += 1
            elif x == 'Strongly Agree':
                q5 += 1
        for x in question6:
            if x == 'Agree':
                q6 += 1
            elif x == 'Strongly Agree':
                q6 += 1
        for x in question7:
            if x == 'Agree':
                q7 += 1
            elif x == 'Strongly Agree':
                q7 += 1
        for x in question8:
            if x == 'Agree':
                q8 += 1
            elif x == 'Strongly Agree':
                q8 += 1
        for x in question9:
            if x == 'Agree':
                q9 += 1
            elif x == 'Strongly Agree':
                q9 += 1
        wb.save("C:\\Users\\short\\Desktop\\JobSeek1.xlsx")
    except:
        print("An error occured while counting responses!")
    else:
        print("Success while counting responses.")

# Function outputs survey data on first Tab
def data_output():
    global surveys
    surveys = 0
    surveys = ss + txk + paris + mtp
    print('\n')
    print("Total Office Counts")
    print('Office count for SS is: ', ss)
    print('Office count for Paris is: ', paris)
    print('Office count for Texarkana is: ', txk)
    print('Office count for Mt Pleasant is: ', mtp)
    print('\n')
    print('The total number of surveys were: ', surveys)
    print('\n')
    print('Question 1 positive responses: ', q1)
    print('Question 2 positive responses: ', q2)
    print('Question 3 positive responses: ', q3)
    print('Question 4 positive responses: ', q4)
    print('Question 5 positive responses: ', q5)
    print('Question 6 positive responses: ', q6)
    print('Question 7 positive responses: ', q7)
    print('Question 8 positive responses: ', q8)
    print('Question 9 positive responses: ', q9)
    print(openpyxl.utils.datetime.from_excel(41250, offset=2415018.5))

# Functions outputs per office numbers onto Survey Tab.
def survey_tab():
    wb = load_workbook("C:\\Users\\short\\Desktop\\JobSeek1.xlsx")
    wb.create_sheet(index=1, title='Survey')
    stab = wb['Survey']
    stab['A1'] = 'Total Office Counts'
    stab['A2'] = 'Mt Pleasant'
    stab['A3'] = 'Paris'
    stab['A4'] = 'Texarkana'
    stab['A5'] = 'Sulfur Springs'
    stab['A6'] = 'Total Surveys'
    stab['B2'] = mtp
    stab['B3'] = paris
    stab['B4'] = txk
    stab['B5'] = ss
    stab['B6'] = surveys
    stab['A8'] = 'Questions'
    stab['A9'] = 'Question1'
    stab['A10'] = 'Question2'
    stab['A11'] = 'Question3'
    stab['A12'] = 'Question4'
    stab['A13'] = 'Question5'
    stab['A14'] = 'Question6'
    stab['A15'] = 'Question7'
    stab['A16'] = 'Question8'
    stab['A17'] = 'Question9'
    stab['B9'] = q1
    stab['B10'] = q2
    stab['B11'] = q3
    stab['B12'] = q4
    stab['B13'] = q5
    stab['B14'] = q6
    stab['B15'] = q7
    stab['B16'] = q8
    stab['B17'] = q9
    stab['B8'] = 'Positives'
    stab['C8'] = 'Total'
    stab['C9'] = surveys
    stab['C10'] = surveys
    stab['C11'] = surveys
    stab['C12'] = surveys
    stab['C13'] = surveys
    stab['C14'] = surveys
    stab['C15'] = surveys
    stab['C16'] = surveys
    stab['C17'] = surveys
    stab['D8'] = 'Negitives'
    stab['D9'] = surveys - q1
    stab['D10'] = surveys - q2
    stab['D11'] = surveys - q3
    stab['D12'] = surveys - q4
    stab['D13'] = surveys - q5
    stab['D14'] = surveys - q6
    stab['D15'] = surveys - q7
    stab['D16'] = surveys - q8
    stab['D17'] = surveys - q9
    wb.save("C:\\Users\\short\\Desktop\\JobSeek1.xlsx")

# Main execution
if __name__ == '__main__':
    del_extra_stuff()
    count_offices()
    question_tally()
    data_output()
    survey_tab()
